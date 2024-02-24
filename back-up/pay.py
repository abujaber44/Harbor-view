import openpyxl
import pandas as pd
import subprocess



# Load the workbook
wb = openpyxl.load_workbook('Daily Sheet.xlsx', data_only=True)

# Create an empty list to store the results
results = []

# Iterate through the first 14 sheets
for sheet in wb.worksheets[:14]:
    # Get the result of the formula in B1
    day = sheet['B1'].value
    
    # Iterate through the cells in M4:M47
    for row in sheet['M4:M47']:
        for cell in row:
            if cell.value is not None:
                if isinstance(cell.value, int) and cell.value < 0:
                    # Copy the values from columns C and M
                    driver = sheet.cell(row=cell.row, column=3).value
                    notes = sheet.cell(row=cell.row, column=14).value
                    # Append the results to the list
                    results.append([day, driver, -1*(cell.value), notes])

# Create a new workbook to store the results
result_wb = openpyxl.Workbook()
result_sheet = result_wb.active

# Write the header row to the new sheet
result_sheet.append(['Day', 'Driver', 'Amount', 'Notes'])

# Write the results to the new sheet
for result in results:
    result_sheet.append(result)

for row in result_sheet.iter_rows(min_row=2, max_col=1):
    for cell in row:
        # Check if the cell is not empty
        if cell.value is not None:
            # Change the data format of the cell
            cell.number_format = 'MM/DD'

# Save the result workbook
result_wb.save('output.xlsx')

print('Output has been saved')
print('Starting express server...')

subprocess.run(["node", "app.js"])


# # Load your excel file into a pandas DataFrame
# df = pd.read_excel("output.xlsx")

# # Convert the desired column to a string format
# df['Day'] = df['Day'].dt.strftime('%m/%d')

# # Save the changes to the same file or to a new file
# df.to_excel("output.xlsx", index=False)
