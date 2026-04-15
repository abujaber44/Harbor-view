import json
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta

try:
    import openpyxl
except ModuleNotFoundError as exc:
    print(
        json.dumps(
            {
                "ok": False,
                "code": "DEPENDENCY_MISSING",
                "message": "Python package openpyxl is missing.",
                "details": str(exc),
            }
        ),
        file=sys.stderr,
    )
    sys.exit(1)


CONFIG = {
    "day_to_sheets": {
        "Monday": ["Mon AM", "Mon PM"],
        "Tuesday": ["Tues AM", "Tues PM"],
        "Wednesday": ["Wed AM", "Wed PM"],
        "Thursday": ["Thurs AM", "Thurs PM"],
        "Friday": ["Fri AM", "Fri PM"],
        "Saturday": ["Sat AM", "Sat PM"],
        "Sunday": ["Sun AM", "Sun PM"],
    },
    "day_order": [
        "Monday",
        "Tuesday",
        "Wednesday",
        "Thursday",
        "Friday",
        "Saturday",
        "Sunday",
    ],
    "cells": {
        "day_cell": "B1",
        "min_row": 4,
        "max_row": 47,
        "driver_col": 3,
        "amount_col": 13,
        "adj_col": 14,
        "notes_col": 15,
    },
    "output_headers": ["Day", "Driver", "Amount", "Adj", "Notes", "Driver2", "Sum"],
}


class PayrollError(Exception):
    def __init__(self, code, message, details=None):
        self.code = code
        self.message = message
        self.details = details
        super().__init__(message)


DAY_FORMULA_RE = re.compile(
    r"^=\s*(?:'([^']+)'|([A-Za-z0-9_ ]+))\s*!\s*\$?B\$?1\s*([+-]\s*\d+)?\s*$",
    re.IGNORECASE,
)


def parse_payload(raw_payload):
    try:
        payload = json.loads(raw_payload or "{}")
    except json.JSONDecodeError as exc:
        raise PayrollError("INPUT_INVALID", "Invalid request payload.", str(exc))

    from_day = payload.get("fromDay")
    to_day = payload.get("toDay")

    if not from_day or not to_day:
        raise PayrollError("INPUT_INVALID", "Both fromDay and toDay are required.")

    return from_day, to_day


def build_sheets_to_process(from_day, to_day, config=CONFIG):
    day_order = config["day_order"]
    day_to_sheets = config["day_to_sheets"]

    if from_day not in day_order or to_day not in day_order:
        raise PayrollError("INPUT_INVALID", "Days must be valid weekday names.")

    from_idx = day_order.index(from_day)
    to_idx = day_order.index(to_day)

    if from_idx > to_idx:
        raise PayrollError("INPUT_INVALID", "fromDay must be earlier than or equal to toDay.")

    sheet_names = []
    for day in day_order[from_idx : to_idx + 1]:
        sheet_names.extend(day_to_sheets[day])

    return sheet_names


def coerce_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value or "").strip()
    if not text:
        return None

    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    try:
        parsed = datetime.strptime(text, "%m/%d")
        return date(datetime.now().year, parsed.month, parsed.day)
    except ValueError:
        return None


def sheet_to_day_map(config=CONFIG):
    mapping = {}
    for day_name, sheet_names in config["day_to_sheets"].items():
        for sheet_name in sheet_names:
            mapping[sheet_name] = day_name
    return mapping


def resolve_sheet_day_labels(workbook_values, workbook_formulas, sheet_names, config=CONFIG):
    cells = config["cells"]
    day_order = config["day_order"]
    sheet_day_lookup = sheet_to_day_map(config)
    resolved_dates = {}

    def resolve_from_formula(sheet_name, stack):
        if sheet_name in resolved_dates:
            return resolved_dates[sheet_name]
        if sheet_name not in workbook_values.sheetnames:
            return None
        if sheet_name in stack:
            return None

        value_sheet = workbook_values[sheet_name]
        raw_value = value_sheet[cells["day_cell"]].value
        parsed = coerce_date(raw_value)
        if parsed:
            resolved_dates[sheet_name] = parsed
            return parsed

        formula_sheet = workbook_formulas[sheet_name]
        formula_value = formula_sheet[cells["day_cell"]].value
        formula_text = str(formula_value or "").strip()
        match = DAY_FORMULA_RE.match(formula_text)
        if not match:
            return None

        ref_sheet = (match.group(1) or match.group(2) or "").strip()
        offset_text = match.group(3) or ""
        offset_days = int(offset_text.replace(" ", "")) if offset_text else 0

        parent = resolve_from_formula(ref_sheet, stack | {sheet_name})
        if not parent:
            return None

        resolved = parent + timedelta(days=offset_days)
        resolved_dates[sheet_name] = resolved
        return resolved

    for sheet_name in sheet_names:
        resolve_from_formula(sheet_name, set())

    anchor = None
    for sheet_name in sheet_names:
        anchor_date = resolved_dates.get(sheet_name)
        anchor_day_name = sheet_day_lookup.get(sheet_name)
        if anchor_date and anchor_day_name in day_order:
            anchor = (anchor_day_name, anchor_date)
            break

    if anchor:
        anchor_day_name, anchor_date = anchor
        anchor_idx = day_order.index(anchor_day_name)
        for sheet_name in sheet_names:
            if sheet_name in resolved_dates:
                continue
            day_name = sheet_day_lookup.get(sheet_name)
            if day_name not in day_order:
                continue
            day_idx = day_order.index(day_name)
            resolved_dates[sheet_name] = anchor_date + timedelta(days=day_idx - anchor_idx)

    labels = {}
    for sheet_name in sheet_names:
        resolved = resolved_dates.get(sheet_name)
        if resolved:
            labels[sheet_name] = resolved.strftime("%m/%d")
            continue

        fallback_day = sheet_day_lookup.get(sheet_name)
        labels[sheet_name] = fallback_day or sheet_name

    return labels


def extract_rows(workbook_path, sheet_names, config=CONFIG):
    try:
        workbook = openpyxl.load_workbook(workbook_path, data_only=True)
        workbook_formulas = openpyxl.load_workbook(workbook_path, data_only=False)
    except FileNotFoundError as exc:
        raise PayrollError("WORKBOOK_MISSING", f"Workbook not found at {workbook_path}.", str(exc))
    except OSError as exc:
        raise PayrollError("PROCESS_FAILED", "Failed to open workbook.", str(exc))

    records = []
    cells = config["cells"]
    day_labels = resolve_sheet_day_labels(workbook, workbook_formulas, sheet_names, config)

    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            raise PayrollError("SHEET_NOT_FOUND", f"Worksheet '{sheet_name}' was not found.")

        sheet = workbook[sheet_name]
        day = day_labels.get(sheet_name)

        for row_idx in range(cells["min_row"], cells["max_row"] + 1):
            amount = sheet.cell(row=row_idx, column=cells["amount_col"]).value
            if isinstance(amount, (int, float)) and amount < 0:
                driver = sheet.cell(row=row_idx, column=cells["driver_col"]).value
                adj = sheet.cell(row=row_idx, column=cells["adj_col"]).value
                notes = sheet.cell(row=row_idx, column=cells["notes_col"]).value

                records.append([
                    day,
                    driver,
                    -1 * amount,
                    adj if isinstance(adj, (int, float)) else 0,
                    notes or "",
                ])

    return records


def write_output(records, output_path, config=CONFIG):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.append(config["output_headers"])
    for row in records:
        sheet.append(row)

    sums = defaultdict(float)
    for record in records:
        driver = record[1]
        amount = record[2]
        if driver is not None:
            sums[driver] += amount

    for row_idx, (driver, total) in enumerate(sums.items(), start=2):
        sheet.cell(row=row_idx, column=6).value = driver
        sheet.cell(row=row_idx, column=7).value = total

    for row_idx in range(2, sheet.max_row + 1):
        day_cell = sheet.cell(row=row_idx, column=1)
        if day_cell.value is not None and not isinstance(day_cell.value, str) and hasattr(day_cell.value, "strftime"):
            day_cell.value = day_cell.value.strftime("%m/%d")

    workbook.save(output_path)


def run(from_day, to_day, workbook_path, output_path):
    sheet_names = build_sheets_to_process(from_day, to_day)
    records = extract_rows(workbook_path, sheet_names)
    write_output(records, output_path)

    warnings = []
    if not records:
        warnings.append("No negative values were found for the selected range.")

    return {
        "ok": True,
        "outputFile": os.path.basename(output_path),
        "rowsWritten": len(records),
        "warnings": warnings,
    }


def main():
    workbook_path = os.getenv("PAYROLL_WORKBOOK", "Daily Sheet.xlsx")
    output_path = os.getenv("PAYROLL_OUTPUT", "output.xlsx")

    try:
        from_day, to_day = parse_payload(sys.stdin.read())
        result = run(from_day, to_day, workbook_path, output_path)
        print(json.dumps(result))
        return 0
    except PayrollError as exc:
        print(
            json.dumps(
                {
                    "ok": False,
                    "code": exc.code,
                    "message": exc.message,
                    "details": exc.details,
                }
            ),
            file=sys.stderr,
        )
        return 1
    except Exception as exc:
        print(
            json.dumps(
                {
                    "ok": False,
                    "code": "PROCESS_FAILED",
                    "message": "Unexpected payroll processing error.",
                    "details": str(exc),
                }
            ),
            file=sys.stderr,
        )
        return 1


if __name__ == "__main__":
    sys.exit(main())
