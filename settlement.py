import json
import os
import shutil
import sys
from copy import deepcopy
from datetime import date, datetime

try:
    import openpyxl
except ModuleNotFoundError as exc:
    print(
        json.dumps(
            {
                "ok": False,
                "code": "DEPENDENCY_MISSING",
                "message": "Python package openpyxl is missing.",
                "details": str(exc),
            }
        ),
        file=sys.stderr,
    )
    sys.exit(1)


CONFIG = {
    "day_order": [
        "Monday",
        "Tuesday",
        "Wednesday",
        "Thursday",
        "Friday",
        "Saturday",
        "Sunday",
    ],
    "day_to_sheets": {
        "Monday": [("AM", "Mon AM"), ("PM", "Mon PM")],
        "Tuesday": [("AM", "Tues AM"), ("PM", "Tues PM")],
        "Wednesday": [("AM", "Wed AM"), ("PM", "Wed PM")],
        "Thursday": [("AM", "Thurs AM"), ("PM", "Thurs PM")],
        "Friday": [("AM", "Fri AM"), ("PM", "Fri PM")],
        "Saturday": [("AM", "Sat AM"), ("PM", "Sat PM")],
        "Sunday": [("AM", "Sun AM"), ("PM", "Sun PM")],
    },
    "cells": {
        "date_cell": "B1",
        "min_row": 4,
        "max_row": 47,
        "driver_col": "C",
        "include_flag_col": "D",
        "ntd_col": "L",
        "cash_col": "M",
        "adj_col": "N",
        "notes_col": "O",
    },
}


class SettlementError(Exception):
    def __init__(self, code, message, details=None):
        self.code = code
        self.message = message
        self.details = details
        super().__init__(message)


def normalize_driver_key(value):
    return " ".join(str(value or "").split()).strip().lower()


def normalize_driver_display(value):
    return " ".join(str(value or "").split()).strip()


def normalize_money_value(value):
    if isinstance(value, (int, float)):
        parsed = float(value)
        return parsed if parsed == parsed and parsed not in (float("inf"), float("-inf")) else 0.0

    text = str(value or "").strip()
    if not text:
        return 0.0

    is_parentheses_negative = text.startswith("(") and text.endswith(")")
    stripped = text.replace(",", "").replace("$", "").replace(" ", "").replace("(", "").replace(")", "")

    try:
        parsed = float(stripped)
    except ValueError:
        return 0.0

    return (parsed * -1.0) if is_parentheses_negative else parsed


def round_balance_by_rule(value):
    amount = normalize_money_value(value)
    sign = -1 if amount < 0 else 1
    absolute = abs(amount)
    whole = int(absolute)
    fraction = absolute - whole
    rounded_absolute = whole + 1 if fraction > 0.5 else whole
    return sign * rounded_absolute


def amount_equal(a, b):
    return abs(float(a) - float(b)) <= 1e-9


def clean_amount(value):
    rounded = round(float(value), 6)
    if abs(rounded - int(rounded)) <= 1e-9:
        return int(rounded)
    return rounded


def format_currency_compact(value):
    rounded = round(float(value), 2)
    fixed = f"{rounded:.2f}"
    if fixed.endswith(".00"):
        return fixed[:-3]
    if fixed.endswith("0"):
        return fixed[:-1]
    return fixed


def to_month_day_label(raw_value, fallback_label):
    if isinstance(raw_value, datetime):
        return f"{raw_value.month}/{raw_value.day}"
    if isinstance(raw_value, date):
        return f"{raw_value.month}/{raw_value.day}"

    text = str(raw_value or "").strip()
    if not text:
        return fallback_label
    if text.startswith("=") or "!" in text:
        return fallback_label

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return f"{parsed.month}/{parsed.day}"
        except ValueError:
            continue

    return text


def to_service_date_value(raw_value, fallback_label):
    if isinstance(raw_value, datetime):
        return raw_value.strftime("%Y-%m-%d")
    if isinstance(raw_value, date):
        return raw_value.strftime("%Y-%m-%d")

    text = str(raw_value or "").strip()
    if not text:
        return fallback_label
    if text.startswith("=") or "!" in text:
        return fallback_label

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            continue

    return text


def join_date_labels(labels):
    unique = []
    seen = set()
    for label in labels:
        if label in seen:
            continue
        unique.append(label)
        seen.add(label)

    if len(unique) == 0:
        return ""
    if len(unique) == 1:
        return unique[0]
    if len(unique) == 2:
        return f"{unique[0]} & {unique[1]}"
    return f"{', '.join(unique[:-1])} & {unique[-1]}"


def append_note(existing, addition):
    if not addition:
        return existing
    if not existing:
        return addition
    return f"{existing} | {addition}"


def period_rank(period):
    return 0 if period == "AM" else 1


def compare_row_key(row):
    return (row["day_index"], period_rank(row["period"]), row["row_number"])


def parse_payload(raw_payload):
    try:
        payload = json.loads(raw_payload or "{}")
    except json.JSONDecodeError as exc:
        raise SettlementError("INPUT_INVALID", "Invalid request payload.", str(exc))

    from_day = payload.get("fromDay")
    to_day = payload.get("toDay")

    if not from_day or not to_day:
        raise SettlementError("INPUT_INVALID", "Both fromDay and toDay are required.")

    apply = bool(payload.get("apply", False))
    return from_day, to_day, apply


def build_day_slice(from_day, to_day, config=CONFIG):
    day_order = config["day_order"]
    if from_day not in day_order or to_day not in day_order:
        raise SettlementError("INPUT_INVALID", "Days must be valid weekday names.")

    from_idx = day_order.index(from_day)
    to_idx = day_order.index(to_day)

    if from_idx > to_idx:
        raise SettlementError("INPUT_INVALID", "fromDay must be earlier than or equal to toDay.")

    return from_idx, to_idx


def load_workbook(path_value, data_only=False):
    try:
        return openpyxl.load_workbook(path_value, data_only=data_only)
    except FileNotFoundError as exc:
        raise SettlementError("WORKBOOK_MISSING", f"Workbook not found at {path_value}.", str(exc))
    except OSError as exc:
        raise SettlementError("PROCESS_FAILED", "Failed to open workbook.", str(exc))


def is_non_empty(value):
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def extract_numeric_cell(sheet_values, sheet_edit, cell_ref):
    value_primary = sheet_values[cell_ref].value
    value_fallback = sheet_edit[cell_ref].value

    if isinstance(value_primary, (int, float)):
        return normalize_money_value(value_primary)

    if isinstance(value_fallback, (int, float)):
        return normalize_money_value(value_fallback)

    return normalize_money_value(value_primary)


def collect_rows(workbook_edit, workbook_values, from_idx, to_idx, config=CONFIG):
    rows = []
    day_order = config["day_order"]
    cells = config["cells"]

    for day_index in range(from_idx, to_idx + 1):
        day_name = day_order[day_index]
        for period, sheet_name in config["day_to_sheets"][day_name]:
            if sheet_name not in workbook_edit.sheetnames:
                raise SettlementError("SHEET_NOT_FOUND", f"Worksheet '{sheet_name}' was not found.")

            sheet_edit = workbook_edit[sheet_name]
            sheet_values = workbook_values[sheet_name] if sheet_name in workbook_values.sheetnames else sheet_edit
            service_label = to_month_day_label(sheet_values[cells["date_cell"]].value, day_name)
            service_date = to_service_date_value(sheet_values[cells["date_cell"]].value, day_name)

            for row_number in range(cells["min_row"], cells["max_row"] + 1):
                driver_raw = sheet_edit[f"{cells['driver_col']}{row_number}"].value
                driver_display = normalize_driver_display(driver_raw)
                driver_key = normalize_driver_key(driver_raw)

                if not driver_key:
                    continue

                include_flag = sheet_edit[f"{cells['include_flag_col']}{row_number}"].value
                if not is_non_empty(include_flag):
                    continue

                ntd_cell = f"{cells['ntd_col']}{row_number}"
                cash_cell = f"{cells['cash_col']}{row_number}"
                adj_cell = f"{cells['adj_col']}{row_number}"
                notes_cell = f"{cells['notes_col']}{row_number}"

                ntd = extract_numeric_cell(sheet_values, sheet_edit, ntd_cell)
                cash = extract_numeric_cell(sheet_values, sheet_edit, cash_cell)
                adj_raw = sheet_edit[adj_cell].value
                adj = normalize_money_value(adj_raw)
                notes = str(sheet_edit[notes_cell].value or "").strip()
                adj_has_value = is_non_empty(adj_raw)

                balance = round_balance_by_rule(cash - ntd)
                rows.append(
                    {
                        "id": f"{sheet_name}:{row_number}",
                        "driver": driver_display,
                        "driver_key": driver_key,
                        "day": day_name,
                        "day_index": day_index,
                        "period": period,
                        "sheet_name": sheet_name,
                        "service_label": service_label,
                        "service_date": service_date,
                        "row_number": row_number,
                        "ntd": ntd,
                        "cash_before": cash,
                        "cash_after": cash,
                        "adj_before": adj,
                        "adj_after": adj,
                        "adj_has_value": adj_has_value,
                        "notes_before": notes,
                        "notes_after": notes,
                        "balance_before": balance,
                        "balance_after": balance,
                        "short_before": 0.0,
                        "short_after": 0.0,
                        "short_authoritative": False,
                        "force_write_zero_adj": False,
                    }
                )

    return sorted(rows, key=compare_row_key)


def allocate_settlements(driver_key, source_rows, target_rows):
    sources = sorted(
        [{"remaining": row["available"], **row} for row in source_rows if row["available"] > 0],
        key=compare_row_key,
    )
    targets = sorted(
        [{"remaining": row["needed"], **row} for row in target_rows if row["needed"] > 0],
        key=compare_row_key,
    )

    allocations = []
    total_deducted = 0.0
    source_index = 0

    for target in targets:
        while target["remaining"] > 0 and source_index < len(sources):
            source = sources[source_index]
            if source["remaining"] <= 0:
                source_index += 1
                continue

            amount = min(target["remaining"], source["remaining"])
            if amount <= 0:
                source_index += 1
                continue

            allocations.append(
                {
                    "driverKey": driver_key,
                    "sourceRowId": source["row_id"],
                    "targetRowId": target["row_id"],
                    "sourceDay": source["day"],
                    "sourcePeriod": source["period"],
                    "targetDay": target["day"],
                    "targetPeriod": target["period"],
                    "targetServiceLabel": target["service_label"],
                    "amount": amount,
                }
            )

            source["remaining"] -= amount
            target["remaining"] -= amount
            total_deducted += amount

            if source["remaining"] <= 0:
                source_index += 1

    unresolved = []
    for target in targets:
        if target["remaining"] > 0:
            unresolved.append(
                {
                    "targetRowId": target["row_id"],
                    "targetDay": target["day"],
                    "targetPeriod": target["period"],
                    "remaining": target["remaining"],
                }
            )

    return {
        "allocations": allocations,
        "unresolvedTargets": unresolved,
        "totalDeducted": total_deducted,
    }


def run_settlement(from_day, to_day, workbook_path, should_apply, backup_dir, config=CONFIG):
    from_idx, to_idx = build_day_slice(from_day, to_day, config=config)
    workbook = load_workbook(workbook_path, data_only=False)
    workbook_values = load_workbook(workbook_path, data_only=True)
    rows = collect_rows(workbook, workbook_values, from_idx, to_idx, config=config)

    mutable_rows = deepcopy(rows)
    row_by_id = {row["id"]: row for row in mutable_rows}
    grouped = {}
    for row in mutable_rows:
        grouped.setdefault(row["driver_key"], []).append(row)

    warnings = []
    drivers_processed = 0
    total_deducted = 0.0
    unresolved_shortage_total = 0.0

    for driver_key, driver_rows in grouped.items():
        source_rows = []
        target_rows = []

        for row in driver_rows:
            cash = row["cash_after"]
            ntd = row["ntd"]
            adj = row["adj_after"]
            raw_short = max(0.0, ntd - cash)
            rounded_short = max(0.0, round_balance_by_rule(raw_short))

            if cash < 0:
                source_capacity = abs(cash)
                already_consumed = max(0.0, -adj)
                available = max(0.0, source_capacity - already_consumed)
                if available > 0:
                    source_rows.append(
                        {
                            "row_id": row["id"],
                            "day": row["day"],
                            "day_index": row["day_index"],
                            "period": row["period"],
                            "row_number": row["row_number"],
                            "available": available,
                        }
                    )

            if cash >= 0:
                needed = max(0.0, adj) if row["adj_has_value"] else rounded_short
                if needed > 0:
                    row["short_before"] = needed
                    row["short_after"] = needed
                    row["short_authoritative"] = True
                    if not row["adj_has_value"]:
                        # Persist an explicit zero in N when fully cleared so reruns stay idempotent.
                        row["force_write_zero_adj"] = True

                    target_rows.append(
                        {
                            "row_id": row["id"],
                            "day": row["day"],
                            "day_index": row["day_index"],
                            "period": row["period"],
                            "row_number": row["row_number"],
                            "needed": needed,
                            "service_label": row["service_label"],
                        }
                    )

        if len(target_rows) == 0:
            continue

        drivers_processed += 1
        allocation_result = allocate_settlements(driver_key, source_rows, target_rows)
        total_deducted += allocation_result["totalDeducted"]

        source_note_map = {}
        for allocation in allocation_result["allocations"]:
            source_row = row_by_id.get(allocation["sourceRowId"])
            target_row = row_by_id.get(allocation["targetRowId"])
            if source_row is None or target_row is None:
                continue

            # Target cash is increased when settlement covers shortage.
            # Source negative cash remains unchanged.
            target_row["cash_after"] = target_row["cash_after"] + allocation["amount"]
            target_row["short_after"] = max(0.0, target_row["short_after"] - allocation["amount"])
            target_row["adj_after"] = target_row["short_after"]
            target_row["balance_after"] = round_balance_by_rule(target_row["cash_after"] - target_row["ntd"])

            source_row["adj_after"] = source_row["adj_after"] - allocation["amount"]
            source_row["balance_after"] = round_balance_by_rule(source_row["cash_after"] - source_row["ntd"])

            source_bucket = source_note_map.setdefault(source_row["id"], {"total": 0.0, "labels": []})
            source_bucket["total"] += allocation["amount"]
            source_bucket["labels"].append(allocation["targetServiceLabel"])

        for source_row_id, source_note in source_note_map.items():
            source_row = row_by_id.get(source_row_id)
            if source_row is None:
                continue

            sorted_labels = sorted(source_note["labels"])
            run_note = f"Deducted ${format_currency_compact(source_note['total'])} for {join_date_labels(sorted_labels)}"
            source_row["notes_after"] = append_note(source_row["notes_after"], run_note)

        if len(allocation_result["unresolvedTargets"]) > 0:
            unresolved_for_driver = sum(item["remaining"] for item in allocation_result["unresolvedTargets"])
            unresolved_shortage_total += unresolved_for_driver
            driver_display = driver_rows[0]["driver"] if len(driver_rows) > 0 else driver_key
            warnings.append(
                f"Unresolved shortage for {driver_display}: ${format_currency_compact(unresolved_for_driver)}."
            )

    changed_rows = []
    for row in mutable_rows:
        if row["short_authoritative"]:
            row["adj_after"] = max(0.0, row["short_after"])

        changed = (
            not amount_equal(row["cash_before"], row["cash_after"])
            or not amount_equal(row["adj_before"], row["adj_after"])
            or row["notes_before"] != row["notes_after"]
            or (row["force_write_zero_adj"] and amount_equal(row["adj_after"], 0) and not row["adj_has_value"])
        )
        row["balance_after"] = round_balance_by_rule(row["cash_after"] - row["ntd"])
        row["changed"] = changed
        if changed:
            changed_rows.append(row)

    grouped_changed = {}
    for row in mutable_rows:
        grouped_changed.setdefault(row["driver_key"], []).append(row)

    drivers_preview = []
    drivers_reviewed = len(grouped.keys())
    for driver_rows in grouped_changed.values():
        sorted_rows = sorted(driver_rows, key=compare_row_key)
        starting_open = sum(
            item["short_before"]
            for item in sorted_rows
            if item["short_before"] > 0
        )
        ending_open = sum(
            item["short_after"]
            for item in sorted_rows
            if item["short_after"] > 0
        )
        total_deducted_driver = sum(
            max(0.0, item["short_before"] - item["short_after"]) for item in sorted_rows
        )

        row_entries = []
        for item in sorted_rows:
            row_entries.append(
                {
                    "serviceDate": item["service_date"],
                    "day": item["day"],
                    "period": item["period"],
                    "serviceLabel": item["service_label"],
                    "sheetName": item["sheet_name"],
                    "rowNumber": item["row_number"],
                    "cashInOutBefore": clean_amount(item["cash_before"]),
                    "cashInOutAfter": clean_amount(item["cash_after"]),
                    "adjBefore": clean_amount(item["adj_before"]),
                    "adjAfter": clean_amount(item["adj_after"]),
                    "balanceBefore": clean_amount(item["short_before"]),
                    "balanceAfter": clean_amount(item["short_after"]),
                    "notesBefore": item["notes_before"],
                    "notesAfter": item["notes_after"],
                    "ignored": False,
                    "changed": bool(item["changed"]),
                }
            )

        if ending_open <= 1e-9:
            outcome = "Settled"
        else:
            outcome = f"Open balance remains: ${format_currency_compact(ending_open)}"
            if total_deducted_driver <= 1e-9:
                outcome = f"No deduction applied. {outcome}"

        if starting_open > 0 or any(entry["changed"] for entry in row_entries):
            drivers_preview.append(
                {
                    "driver": sorted_rows[0]["driver"] if len(sorted_rows) > 0 else "",
                    "startingOpenBalance": clean_amount(starting_open),
                    "endingOpenBalance": clean_amount(ending_open),
                    "totalDeducted": clean_amount(total_deducted_driver),
                    "outcome": outcome,
                    "rows": row_entries,
                }
            )

    drivers_preview.sort(
        key=lambda item: (-float(item["startingOpenBalance"]), str(item["driver"]).lower())
    )

    result = {
        "ok": True,
        "fromDay": from_day,
        "toDay": to_day,
        "driversReviewed": drivers_reviewed,
        "driversWithOpenBalance": len(drivers_preview),
        "driversProcessed": drivers_processed,
        "rowsChanged": len(changed_rows),
        "rowsAdjusted": len(changed_rows),
        "totalDeducted": clean_amount(total_deducted),
        "lockedSessionsSkipped": 0,
        "unresolvedShortageTotal": clean_amount(unresolved_shortage_total),
        "warnings": warnings,
        "drivers": drivers_preview,
    }

    if should_apply:
        os.makedirs(backup_dir, exist_ok=True)
        base_name = os.path.splitext(os.path.basename(workbook_path))[0]
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup_path = os.path.join(backup_dir, f"{base_name}.{timestamp}.xlsx")
        shutil.copy2(workbook_path, backup_path)

        cells = config["cells"]
        for row in changed_rows:
            sheet = workbook[row["sheet_name"]]
            sheet[f"{cells['cash_col']}{row['row_number']}"] = clean_amount(row["cash_after"])
            sheet[f"{cells['adj_col']}{row['row_number']}"] = clean_amount(row["adj_after"])
            sheet[f"{cells['notes_col']}{row['row_number']}"] = row["notes_after"]

        workbook.save(workbook_path)
        result["backupFile"] = os.path.basename(backup_path)
        result["workbookFile"] = os.path.basename(workbook_path)

    return result


def main():
    workbook_path = os.getenv("PAYROLL_WORKBOOK", "Daily Sheet.xlsx")
    backup_dir = os.getenv("PAYROLL_BACKUP_DIR", "backups")

    try:
        from_day, to_day, should_apply = parse_payload(sys.stdin.read())
        result = run_settlement(from_day, to_day, workbook_path, should_apply, backup_dir)
        print(json.dumps(result))
        return 0
    except SettlementError as exc:
        print(
            json.dumps(
                {
                    "ok": False,
                    "code": exc.code,
                    "message": exc.message,
                    "details": exc.details,
                }
            ),
            file=sys.stderr,
        )
        return 1
    except Exception as exc:
        print(
            json.dumps(
                {
                    "ok": False,
                    "code": "PROCESS_FAILED",
                    "message": "Unexpected settlement processing error.",
                    "details": str(exc),
                }
            ),
            file=sys.stderr,
        )
        return 1


if __name__ == "__main__":
    sys.exit(main())
