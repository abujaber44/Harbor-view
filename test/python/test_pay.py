import os
import tempfile
import unittest
from datetime import datetime

import openpyxl

from pay import PayrollError, build_sheets_to_process, extract_rows


class PayScriptTests(unittest.TestCase):
    def test_build_sheets_to_process_includes_saturday(self):
        sheets = build_sheets_to_process("Friday", "Saturday")
        self.assertEqual(sheets, ["Fri AM", "Fri PM", "Sat AM", "Sat PM"])

    def test_extract_rows_reads_negative_values(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = os.path.join(tmpdir, "sample.xlsx")
            workbook = openpyxl.Workbook()

            mon_am = workbook.active
            mon_am.title = "Mon AM"
            mon_am["B1"] = datetime(2026, 3, 23)
            mon_am["C4"] = "Driver One"
            mon_am["M4"] = -120
            mon_am["N4"] = -10
            mon_am["O4"] = "Late"

            mon_pm = workbook.create_sheet("Mon PM")
            mon_pm["B1"] = datetime(2026, 3, 23)

            workbook.save(workbook_path)

            rows = extract_rows(workbook_path, ["Mon AM", "Mon PM"])
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0][1], "Driver One")
            self.assertEqual(rows[0][2], 120)
            self.assertEqual(rows[0][3], -10)

    def test_extract_rows_raises_missing_sheet(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = os.path.join(tmpdir, "sample.xlsx")
            workbook = openpyxl.Workbook()
            workbook.active.title = "Mon AM"
            workbook.save(workbook_path)

            with self.assertRaises(PayrollError) as ctx:
                extract_rows(workbook_path, ["Mon AM", "Mon PM"])

            self.assertEqual(ctx.exception.code, "SHEET_NOT_FOUND")

    def test_extract_rows_resolves_day_from_formula_cell(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = os.path.join(tmpdir, "sample.xlsx")
            workbook = openpyxl.Workbook()

            mon_am = workbook.active
            mon_am.title = "Mon AM"
            mon_am["B1"] = datetime(2026, 4, 6)

            tues_am = workbook.create_sheet("Tues AM")
            tues_am["B1"] = "='Mon AM'!B1+1"
            tues_am["C4"] = "Driver Two"
            tues_am["M4"] = -75

            workbook.save(workbook_path)

            rows = extract_rows(workbook_path, ["Tues AM"])
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0][0], "04/07")
            self.assertEqual(rows[0][1], "Driver Two")
            self.assertEqual(rows[0][2], 75)


if __name__ == "__main__":
    unittest.main()
