import os
import tempfile
import unittest
from datetime import datetime

import openpyxl

from settlement import allocate_settlements, run_settlement


ALL_SHEETS = [
    "Mon AM",
    "Mon PM",
    "Tues AM",
    "Tues PM",
    "Wed AM",
    "Wed PM",
    "Thurs AM",
    "Thurs PM",
    "Fri AM",
    "Fri PM",
    "Sat AM",
    "Sat PM",
    "Sun AM",
    "Sun PM",
]

SHEET_DATES = {
    "Mon AM": datetime(2026, 3, 30),
    "Mon PM": datetime(2026, 3, 30),
    "Tues AM": datetime(2026, 3, 31),
    "Tues PM": datetime(2026, 3, 31),
    "Wed AM": datetime(2026, 4, 1),
    "Wed PM": datetime(2026, 4, 1),
    "Thurs AM": datetime(2026, 4, 2),
    "Thurs PM": datetime(2026, 4, 2),
    "Fri AM": datetime(2026, 4, 3),
    "Fri PM": datetime(2026, 4, 3),
    "Sat AM": datetime(2026, 4, 4),
    "Sat PM": datetime(2026, 4, 4),
    "Sun AM": datetime(2026, 4, 5),
    "Sun PM": datetime(2026, 4, 5),
}


def create_workbook(file_path, rows_by_sheet):
    workbook = openpyxl.Workbook()
    first_sheet = workbook.active
    first_sheet.title = ALL_SHEETS[0]

    for sheet_name in ALL_SHEETS:
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
        sheet["B1"] = SHEET_DATES[sheet_name]
        for row_data in rows_by_sheet.get(sheet_name, []):
            row_number = row_data["row"]
            sheet[f"C{row_number}"] = row_data.get("driver", "")
            sheet[f"D{row_number}"] = row_data.get("include", "x")
            sheet[f"L{row_number}"] = row_data.get("ntd", 0)
            sheet[f"M{row_number}"] = row_data.get("cash", 0)
            if "adj" in row_data:
                sheet[f"N{row_number}"] = row_data.get("adj")
            sheet[f"O{row_number}"] = row_data.get("notes", "")

    workbook.save(file_path)


def read_cells(file_path, sheet_name, row_number):
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    sheet = workbook[sheet_name]
    return {
        "cash": sheet[f"M{row_number}"].value,
        "adj": sheet[f"N{row_number}"].value,
        "notes": str(sheet[f"O{row_number}"].value or ""),
    }


class SettlementAllocatorTests(unittest.TestCase):
    def test_single_source_single_target(self):
        result = allocate_settlements(
            "gus",
            [
                {"row_id": "s1", "day": "Monday", "day_index": 0, "period": "AM", "row_number": 4, "available": 100}
            ],
            [
                {"row_id": "t1", "day": "Wednesday", "day_index": 2, "period": "AM", "row_number": 4, "needed": 70, "service_label": "4/1"}
            ],
        )

        self.assertEqual(result["totalDeducted"], 70)
        self.assertEqual(len(result["allocations"]), 1)
        self.assertEqual(result["allocations"][0]["sourceRowId"], "s1")
        self.assertEqual(result["allocations"][0]["targetRowId"], "t1")
        self.assertEqual(len(result["unresolvedTargets"]), 0)

    def test_multi_source_fill(self):
        result = allocate_settlements(
            "driver",
            [
                {"row_id": "s1", "day": "Monday", "day_index": 0, "period": "AM", "row_number": 4, "available": 40},
                {"row_id": "s2", "day": "Tuesday", "day_index": 1, "period": "AM", "row_number": 4, "available": 60},
            ],
            [
                {"row_id": "t1", "day": "Wednesday", "day_index": 2, "period": "AM", "row_number": 4, "needed": 90, "service_label": "4/1"}
            ],
        )

        self.assertEqual(result["totalDeducted"], 90)
        self.assertEqual(len(result["allocations"]), 2)
        self.assertEqual(result["allocations"][0]["amount"], 40)
        self.assertEqual(result["allocations"][1]["amount"], 50)

    def test_source_spans_multiple_targets(self):
        result = allocate_settlements(
            "driver",
            [
                {"row_id": "s1", "day": "Monday", "day_index": 0, "period": "AM", "row_number": 4, "available": 176}
            ],
            [
                {"row_id": "t1", "day": "Tuesday", "day_index": 1, "period": "AM", "row_number": 4, "needed": 100, "service_label": "3/31"},
                {"row_id": "t2", "day": "Wednesday", "day_index": 2, "period": "AM", "row_number": 4, "needed": 50, "service_label": "4/1"},
            ],
        )

        self.assertEqual(result["totalDeducted"], 150)
        self.assertEqual(len(result["allocations"]), 2)
        self.assertEqual(result["allocations"][0]["targetRowId"], "t1")
        self.assertEqual(result["allocations"][1]["targetRowId"], "t2")

    def test_unresolved_shortage(self):
        result = allocate_settlements(
            "driver",
            [
                {"row_id": "s1", "day": "Monday", "day_index": 0, "period": "AM", "row_number": 4, "available": 30}
            ],
            [
                {"row_id": "t1", "day": "Wednesday", "day_index": 2, "period": "AM", "row_number": 4, "needed": 90, "service_label": "4/1"}
            ],
        )

        self.assertEqual(result["totalDeducted"], 30)
        self.assertEqual(len(result["unresolvedTargets"]), 1)
        self.assertEqual(result["unresolvedTargets"][0]["remaining"], 60)

    def test_fifo_order_day_period_row(self):
        result = allocate_settlements(
            "driver",
            [
                {"row_id": "s1", "day": "Monday", "day_index": 0, "period": "PM", "row_number": 8, "available": 20},
                {"row_id": "s2", "day": "Tuesday", "day_index": 1, "period": "AM", "row_number": 3, "available": 20},
                {"row_id": "s0", "day": "Monday", "day_index": 0, "period": "AM", "row_number": 9, "available": 20},
            ],
            [
                {"row_id": "t1", "day": "Wednesday", "day_index": 2, "period": "AM", "row_number": 4, "needed": 50, "service_label": "4/1"}
            ],
        )

        self.assertEqual([item["sourceRowId"] for item in result["allocations"]], ["s0", "s1", "s2"])


class SettlementWorkbookTests(unittest.TestCase):
    def test_driver_owed_daily_is_not_settled_against_same_rows(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = os.path.join(tmpdir, "Daily Sheet.xlsx")
            backups_dir = os.path.join(tmpdir, "backups")

            create_workbook(
                workbook_path,
                {
                    "Mon AM": [{"row": 7, "driver": "Obaydou 326", "ntd": 0, "cash": -113}],
                    "Tues AM": [{"row": 7, "driver": "Obaydou 326", "ntd": 0, "cash": -86}],
                    "Wed AM": [{"row": 7, "driver": "Obaydou 326", "ntd": 0, "cash": -62}],
                    "Thurs AM": [{"row": 7, "driver": "Obaydou 326", "ntd": 0, "cash": -131}],
                    "Fri AM": [{"row": 7, "driver": "Obaydou 326", "ntd": 0, "cash": -35}],
                },
            )

            preview = run_settlement("Monday", "Friday", workbook_path, False, backups_dir)
            self.assertTrue(preview["ok"])
            self.assertEqual(preview["totalDeducted"], 0)
            self.assertEqual(preview["rowsAdjusted"], 0)
            self.assertEqual(preview["driversProcessed"], 0)
            self.assertEqual(preview["driversWithOpenBalance"], 0)

            before_mon = read_cells(workbook_path, "Mon AM", 7)
            apply_result = run_settlement("Monday", "Friday", workbook_path, True, backups_dir)
            self.assertTrue(apply_result["ok"])
            self.assertEqual(apply_result["totalDeducted"], 0)
            self.assertEqual(apply_result["rowsAdjusted"], 0)

            after_mon = read_cells(workbook_path, "Mon AM", 7)
            self.assertEqual(before_mon, after_mon)

    def test_preview_does_not_mutate_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = os.path.join(tmpdir, "Daily Sheet.xlsx")
            backups_dir = os.path.join(tmpdir, "backups")

            create_workbook(
                workbook_path,
                {
                    "Mon AM": [{"row": 4, "driver": "Preview Driver", "ntd": -120, "cash": -120}],
                    "Wed AM": [{"row": 4, "driver": "Preview Driver", "ntd": 80, "cash": 0}],
                },
            )

            before_source = read_cells(workbook_path, "Mon AM", 4)
            before_target = read_cells(workbook_path, "Wed AM", 4)

            result = run_settlement("Monday", "Wednesday", workbook_path, False, backups_dir)
            self.assertTrue(result["ok"])
            self.assertEqual(result["totalDeducted"], 80)
            self.assertEqual(result["rowsAdjusted"], 2)

            after_source = read_cells(workbook_path, "Mon AM", 4)
            after_target = read_cells(workbook_path, "Wed AM", 4)
            self.assertEqual(before_source, after_source)
            self.assertEqual(before_target, after_target)
            self.assertFalse(os.path.exists(backups_dir))

    def test_apply_updates_cells_and_is_idempotent(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = os.path.join(tmpdir, "Daily Sheet.xlsx")
            backups_dir = os.path.join(tmpdir, "backups")

            create_workbook(
                workbook_path,
                {
                    "Mon AM": [{"row": 4, "driver": "Gus", "ntd": -200, "cash": -200, "notes": "manual"}],
                    "Wed AM": [{"row": 4, "driver": "Gus", "ntd": 70, "cash": 0}],
                    "Thurs AM": [{"row": 4, "driver": "Gus", "ntd": 80, "cash": 0}],
                },
            )

            first = run_settlement("Monday", "Thursday", workbook_path, True, backups_dir)
            self.assertTrue(first["ok"])
            self.assertEqual(first["totalDeducted"], 150)
            self.assertEqual(first["unresolvedShortageTotal"], 0)
            self.assertTrue(first.get("backupFile"))
            self.assertTrue(os.path.exists(os.path.join(backups_dir, first["backupFile"])))

            monday = read_cells(workbook_path, "Mon AM", 4)
            wednesday = read_cells(workbook_path, "Wed AM", 4)
            thursday = read_cells(workbook_path, "Thurs AM", 4)

            self.assertEqual(monday["cash"], -200)
            self.assertEqual(monday["adj"], -150)
            self.assertIn("manual | Deducted $150 for 4/1 & 4/2", monday["notes"])
            self.assertEqual(wednesday["cash"], 70)
            self.assertEqual(thursday["cash"], 80)
            self.assertEqual(wednesday["adj"], 0)
            self.assertEqual(thursday["adj"], 0)

            second = run_settlement("Monday", "Thursday", workbook_path, True, backups_dir)
            self.assertTrue(second["ok"])
            self.assertEqual(second["totalDeducted"], 0)
            self.assertEqual(second["rowsAdjusted"], 0)

            monday_after = read_cells(workbook_path, "Mon AM", 4)
            self.assertEqual(monday_after["adj"], -150)
            self.assertEqual(monday_after["notes"].count("Deducted"), 1)

    def test_expanded_range_uses_remaining_capacity(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = os.path.join(tmpdir, "Daily Sheet.xlsx")
            backups_dir = os.path.join(tmpdir, "backups")

            create_workbook(
                workbook_path,
                {
                    "Mon AM": [{"row": 4, "driver": "Carry Driver", "ntd": -75, "cash": -75}],
                    "Tues AM": [{"row": 4, "driver": "Carry Driver", "ntd": 45, "cash": 0}],
                    "Thurs AM": [{"row": 4, "driver": "Carry Driver", "ntd": 50, "cash": 0}],
                },
            )

            first = run_settlement("Monday", "Tuesday", workbook_path, True, backups_dir)
            self.assertEqual(first["totalDeducted"], 45)

            second = run_settlement("Monday", "Thursday", workbook_path, True, backups_dir)
            self.assertEqual(second["totalDeducted"], 30)
            self.assertEqual(second["unresolvedShortageTotal"], 20)

            monday = read_cells(workbook_path, "Mon AM", 4)
            tuesday = read_cells(workbook_path, "Tues AM", 4)
            thursday = read_cells(workbook_path, "Thurs AM", 4)

            self.assertEqual(monday["adj"], -75)
            self.assertEqual(tuesday["cash"], 45)
            self.assertEqual(thursday["cash"], 30)
            self.assertEqual(tuesday["adj"], 0)
            self.assertEqual(thursday["adj"], 20)


if __name__ == "__main__":
    unittest.main()
